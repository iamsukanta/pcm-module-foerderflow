"""Tariff import pipeline (Module PCM — Tariff Registry Import Wizard).

Parses a collective-agreement grid (grade rows × tier columns) from CSV or Excel
into ``salary_tariffs`` rows, previews them against the existing data (overlap
detection, validation) without writing, and commits on confirm.

Source handling:
  CSV / EXCEL  — fully parsed here (csv stdlib / openpyxl).
  IMAGE        — no OCR engine is bundled; the wizard's Image path collects a
                 manually transcribed grid which is posted to /confirm as ``rows``
                 exactly like the CSV path. This service therefore treats IMAGE
                 and MANUAL identically at confirm time (the grid is already
                 structured) — there is no faked OCR confidence.

Amount parsing accepts German (``3.200,00`` / ``€3.200,00``) and international
(``3200.00`` / ``3200``) formats.
"""

from __future__ import annotations

import csv
import io
import re
import secrets
from datetime import date, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.errors import APIError
from app.models.pcm_tariff import SalaryTariff
from app.services.pcm._validate import opt_date, opt_num, parse_date, req_num, req_str
from app.services.pcm.tariff_lookup import assert_fiscal_year_open, assert_window_valid
from app.utils.serialization import decimal_str

_GRADE_HEADERS = {"grade", "entgeltgruppe", "eg", "gruppe", "salary_group"}
_OPEN_END = date.max


def _parse_amount(raw: Any) -> float | None:
    """Parse a EUR amount cell. Returns None for empty cells; raises on garbage."""
    if raw is None:
        return None
    if isinstance(raw, int | float | Decimal):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    s = s.replace("€", "").replace(" ", "").replace(" ", "")
    if "," in s and "." in s:
        # German thousands + decimal: 3.200,00 → 3200.00
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # Single comma = decimal separator: 3200,50 → 3200.50
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError as exc:
        raise APIError(
            422,
            "IMPORT_PARSE_FAILED",
            f"Betrag konnte nicht gelesen werden: '{raw}'.",
        ) from exc


def _normalise_group(raw: str) -> str:
    return re.sub(r"\s+", "", str(raw).strip())


class TariffImportService:
    def __init__(self, db: Session):
        self.db = db

    # ── parsing ───────────────────────────────────────────────────────────────
    def _read_grid(self, content: bytes, filename: str, source: str) -> list[list[Any]]:
        name = (filename or "").lower()
        if source == "EXCEL" or name.endswith((".xlsx", ".xls")):
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
            ws = wb[wb.sheetnames[0]]
            return [list(row) for row in ws.iter_rows(values_only=True)]
        # CSV path — decode tolerantly, sniff delimiter.
        text: str | None = None
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise APIError(500, "IMPORT_PARSE_FAILED", "Datei-Encoding nicht lesbar.")
        first_line = text.splitlines()[0] if text.splitlines() else ""
        delim = ";" if first_line.count(";") > first_line.count(",") else ","
        return list(csv.reader(io.StringIO(text), delimiter=delim))

    def parse(self, content: bytes, filename: str, source: str) -> list[dict[str, Any]]:
        grid = [r for r in self._read_grid(content, filename, source) if any(
            c is not None and str(c).strip() for c in r
        )]
        if not grid:
            raise APIError(422, "IMPORT_PARSE_FAILED", "Keine Daten in der Datei gefunden.")
        header = grid[0]
        # Grade column = first column (or a column whose header names a grade).
        grade_col = 0
        for idx, h in enumerate(header):
            if h is not None and str(h).strip().lower() in _GRADE_HEADERS:
                grade_col = idx
                break
        tier_cols = [i for i in range(len(header)) if i != grade_col]
        rows: list[dict[str, Any]] = []
        for raw_row in grid[1:]:
            if grade_col >= len(raw_row):
                continue
            group = _normalise_group(raw_row[grade_col] or "")
            if not group:
                continue
            for level, col in enumerate(tier_cols, start=1):
                cell = raw_row[col] if col < len(raw_row) else None
                amount = _parse_amount(cell)
                if amount is None:
                    continue
                rows.append(
                    {"salary_group": group, "level": level, "monthly_amount": amount}
                )
        return rows

    # ── overlap finder (shared by preview + confirm trim) ─────────────────────
    def _conflict(
        self,
        org_id: str,
        *,
        tariff_code: str,
        salary_group: str,
        level: int,
        is_proposed: bool,
        valid_from: date,
        valid_to: date | None,
    ) -> SalaryTariff | None:
        new_to = valid_to or _OPEN_END
        existing = (
            self.db.execute(
                select(SalaryTariff).where(
                    SalaryTariff.org_id == org_id,
                    SalaryTariff.tariff_code == tariff_code,
                    SalaryTariff.salary_group == salary_group,
                    SalaryTariff.level == level,
                    SalaryTariff.is_proposed.is_(is_proposed),
                    SalaryTariff.deleted_at.is_(None),
                )
            )
            .scalars()
            .all()
        )
        for row in existing:
            row_to = row.valid_to or _OPEN_END
            if row.valid_from <= new_to and row_to >= valid_from:
                return row
        return None

    # ── preview (no write) ────────────────────────────────────────────────────
    def preview(
        self, org_id: str, *, content: bytes, filename: str, meta: dict[str, Any]
    ) -> dict[str, Any]:
        source = req_str(meta, "source")
        tariff_code = req_str(meta, "tariff_code")
        is_proposed = str(meta.get("is_proposed", "false")).lower() == "true"
        valid_from = parse_date(meta.get("valid_from"), "valid_from")
        valid_to = opt_date(meta.get("valid_to"), "valid_to")
        assert_window_valid(valid_from, valid_to)

        parsed = self.parse(content, filename, source)
        preview_rows: list[dict[str, Any]] = []
        conflicts: list[dict[str, Any]] = []
        valid_count = warning_count = error_count = 0
        for row in parsed:
            status = "valid"
            conflict = None
            if row["monthly_amount"] <= 0:
                status = "error"
                error_count += 1
            else:
                hit = self._conflict(
                    org_id,
                    tariff_code=tariff_code,
                    salary_group=row["salary_group"],
                    level=row["level"],
                    is_proposed=is_proposed,
                    valid_from=valid_from,
                    valid_to=valid_to,
                )
                if hit is not None:
                    status = "warning"
                    warning_count += 1
                    conflict = {
                        "id": hit.id,
                        "valid_from": hit.valid_from.isoformat(),
                        "valid_to": hit.valid_to.isoformat() if hit.valid_to else None,
                        "monthly_amount": decimal_str(hit.monthly_amount),
                    }
                    conflicts.append({**row, "conflict": conflict})
                else:
                    valid_count += 1
            preview_rows.append({**row, "status": status, "conflict": conflict})

        return {
            "import_id": "imp_" + secrets.token_hex(8),
            "tariff_code": tariff_code,
            "row_count": len(parsed),
            "valid_rows": valid_count,
            "warning_rows": warning_count,
            "error_rows": error_count,
            "preview": preview_rows,
            "conflicts": conflicts,
        }

    # ── confirm (write) ───────────────────────────────────────────────────────
    def confirm(self, org_id: str, body: dict[str, Any]) -> dict[str, Any]:
        tariff_code = req_str(body, "tariff_code")
        is_proposed = bool(body.get("is_proposed", False))
        valid_from = parse_date(body.get("valid_from"), "valid_from")
        valid_to = opt_date(body.get("valid_to"), "valid_to")
        standard_hours = req_num(body, "standard_hours")
        bav_rate_pct = opt_num(body, "bav_rate_pct")
        resolution = str(body.get("conflict_resolution", "skip")).lower()
        rows = body.get("rows") or []
        if standard_hours <= 0:
            raise APIError(422, "VALIDATION_STANDARD_HOURS", "standard_hours muss > 0 sein.")
        assert_window_valid(valid_from, valid_to)
        assert_fiscal_year_open(self.db, org_id=org_id, on=valid_from)

        written = skipped = trimmed = 0
        for row in rows:
            group = req_str(row, "salary_group")
            level = int(row["level"])
            amount = req_num(row, "monthly_amount")
            if amount <= 0:
                skipped += 1
                continue
            hit = self._conflict(
                org_id,
                tariff_code=tariff_code,
                salary_group=group,
                level=level,
                is_proposed=is_proposed,
                valid_from=valid_from,
                valid_to=valid_to,
            )
            if hit is not None:
                if resolution == "trim":
                    hit.valid_to = valid_from - timedelta(days=1)
                    trimmed += 1
                else:
                    skipped += 1
                    continue
            self.db.add(
                SalaryTariff(
                    org_id=org_id,
                    tariff_code=tariff_code,
                    salary_group=group,
                    level=level,
                    monthly_amount=Decimal(str(amount)),
                    standard_hours=Decimal(str(standard_hours)),
                    is_proposed=is_proposed,
                    valid_from=valid_from,
                    valid_to=valid_to,
                    bav_rate_pct=(
                        Decimal(str(bav_rate_pct)) if bav_rate_pct is not None else None
                    ),
                )
            )
            written += 1
        self.db.commit()
        return {
            "tariff_code": tariff_code,
            "written": written,
            "skipped": skipped,
            "trimmed": trimmed,
        }
