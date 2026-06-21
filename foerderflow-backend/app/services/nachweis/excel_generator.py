"""Verwendungsnachweis Excel — openpyxl port of lib/nachweis/excel-generator.ts.

Three sheets: Einnahmen & Ausgaben, Belegliste, Soll-Ist-Vergleich. Styles
(fonts, fills, borders, number formats) reproduce the monolith's ExcelJS output.
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EUR_FMT = '#,##0.00 "€"'


def _fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def _style_title(cell) -> None:
    cell.font = Font(bold=True, size=13, color="FFFFFFFF", name="Calibri")
    cell.fill = _fill("FF1F3864")
    cell.alignment = Alignment(vertical="center", horizontal="left")


def _style_info(cell) -> None:
    cell.font = Font(size=9, color="FF595959", italic=True, name="Calibri")
    cell.fill = _fill("FFF2F2F2")
    cell.alignment = Alignment(vertical="center", horizontal="left")


def _style_section(cell) -> None:
    cell.font = Font(bold=True, size=11, color="FFFFFFFF", name="Calibri")
    cell.fill = _fill("FF2E75B6")
    cell.alignment = Alignment(vertical="center", horizontal="left")


def _style_colheader(cell) -> None:
    cell.font = Font(bold=True, size=10, color="FF1F3864", name="Calibri")
    cell.fill = _fill("FFDAE3F3")
    cell.alignment = Alignment(vertical="center", horizontal="center")
    cell.border = Border(bottom=Side(style="thin", color="FF2E75B6"))


def _style_subgroup(cell) -> None:
    cell.font = Font(bold=True, size=10, italic=True, color="FF404040", name="Calibri")
    cell.fill = _fill("FFF2F2F2")
    cell.alignment = Alignment(vertical="center", horizontal="left")


def _style_data(cell, is_amount: bool = False) -> None:
    cell.font = Font(size=10, name="Calibri")
    cell.alignment = Alignment(vertical="center", horizontal="right" if is_amount else "left")


def _style_sum(cell, is_amount: bool = False) -> None:
    cell.font = Font(bold=True, size=10, name="Calibri")
    cell.fill = _fill("FFFFE699")
    cell.alignment = Alignment(vertical="center", horizontal="right" if is_amount else "left")
    cell.border = Border(
        top=Side(style="thin", color="FFBFBFBF"),
        bottom=Side(style="double", color="FF404040"),
    )


def _style_saldo(cell, positive: bool, is_amount: bool = False) -> None:
    cell.font = Font(bold=True, size=10, name="Calibri")
    cell.fill = _fill("FFE2EFDA" if positive else "FFFFC7CE")
    cell.alignment = Alignment(vertical="center", horizontal="right" if is_amount else "left")
    cell.border = Border(
        top=Side(style="medium", color="FF404040"),
        bottom=Side(style="medium", color="FF404040"),
    )


def _eur(cell) -> None:
    cell.number_format = EUR_FMT


def _de_date(iso: str) -> str:
    try:
        d = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return d.strftime("%d.%m.%Y")
    except ValueError:
        return iso


def _build_einnahmen_ausgaben(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Einnahmen & Ausgaben")
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 20
    m = data["massnahme"]
    e = data["einnahmen"]
    row = 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = f"{m['name']} — Verwendungsnachweis {data['fiscal_year']['jahr']}"
    _style_title(c)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = (
        f"{data['org']['name']}  |  Fördergeber: {m['funder_name']}  |  "
        f"Erstellt: {_de_date(data['generated_at'])}"
    )
    _style_info(c)
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = f"Förderquote {m['foerderquote']:.0f}%  |  Laufzeit {m['laufzeit_von']} bis {m['laufzeit_bis']}"
    _style_info(c)
    row += 2

    # E – EINNAHMEN
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = "E – EINNAHMEN"
    _style_section(c)
    row += 1
    ha, hb = ws[f"A{row}"], ws[f"B{row}"]
    ha.value, hb.value = "Einnahme-Art", "Betrag (€)"
    _style_colheader(ha)
    _style_colheader(hb)
    row += 1

    for label, val in [
        ("Eigenmittel (Eigenanteil der Organisation)", e["eigenmittel"]),
        (f"Zuwendung (Fördermittel {m['funder_name']})", e["zuwendung"]),
        ("Sonstige Einnahmen", e["sonstige"]),
    ]:
        a, b = ws[f"A{row}"], ws[f"B{row}"]
        a.value, b.value = label, val
        _style_data(a)
        _style_data(b, True)
        _eur(b)
        row += 1

    summe_einnahmen = e["eigenmittel"] + e["zuwendung"] + e["sonstige"]
    a, b = ws[f"A{row}"], ws[f"B{row}"]
    a.value, b.value = "SUMME EINNAHMEN", summe_einnahmen
    _style_sum(a)
    _style_sum(b, True)
    _eur(b)
    row += 2

    # A – AUSGABEN
    ws.merge_cells(f"A{row}:B{row}")
    c = ws[f"A{row}"]
    c.value = "A – AUSGABEN"
    _style_section(c)
    row += 1
    ha, hb = ws[f"A{row}"], ws[f"B{row}"]
    ha.value, hb.value = "Kostenart / Bezeichnung", "Betrag (€)"
    _style_colheader(ha)
    _style_colheader(hb)
    row += 1

    personal = [a for a in data["ausgaben"] if a["ist_personal"]]
    sach = [a for a in data["ausgaben"] if not a["ist_personal"]]

    def _group(title: str, items: list[dict], empty_label: str) -> float:
        nonlocal row
        ws.merge_cells(f"A{row}:B{row}")
        g = ws[f"A{row}"]
        g.value = title
        _style_subgroup(g)
        row += 1
        total = 0.0
        if not items:
            a, b = ws[f"A{row}"], ws[f"B{row}"]
            a.value, b.value = empty_label, 0
            _style_data(a)
            _style_data(b, True)
            _eur(b)
            a.font = Font(size=10, name="Calibri", italic=True, color="FF808080")
            row += 1
        else:
            for it in items:
                label = (
                    f"{it['kostenart']} ({it['anteil_beschreibung']})"
                    if it["anteil_beschreibung"]
                    else it["kostenart"]
                )
                a, b = ws[f"A{row}"], ws[f"B{row}"]
                a.value, b.value = label, it["betrag_foerderfahig"]
                _style_data(a)
                _style_data(b, True)
                _eur(b)
                total += it["betrag_foerderfahig"]
                row += 1
        return total

    sum_personal = _group("I. Personalausgaben", personal, "(keine Personalausgaben)")
    a, b = ws[f"A{row}"], ws[f"B{row}"]
    a.value, b.value = "Zwischensumme Personalausgaben", sum_personal
    _style_sum(a)
    _style_sum(b, True)
    _eur(b)
    row += 2

    sum_sach = _group("II. Sachausgaben", sach, "(keine Sachausgaben)")
    a, b = ws[f"A{row}"], ws[f"B{row}"]
    a.value, b.value = "Zwischensumme Sachausgaben", sum_sach
    _style_sum(a)
    _style_sum(b, True)
    _eur(b)
    row += 2

    summe_ausgaben = sum_personal + sum_sach
    a, b = ws[f"A{row}"], ws[f"B{row}"]
    a.value, b.value = "SUMME AUSGABEN", summe_ausgaben
    _style_sum(a)
    _style_sum(b, True)
    _eur(b)
    row += 2

    saldo = summe_einnahmen - summe_ausgaben
    a, b = ws[f"A{row}"], ws[f"B{row}"]
    a.value, b.value = "SALDO (Einnahmen – Ausgaben)", saldo
    _style_saldo(a, saldo >= 0)
    _style_saldo(b, saldo >= 0, True)
    _eur(b)


def _build_belegliste(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Belegliste")
    widths = [6, 12, 22, 32, 14, 18, 16, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"Belegliste — {data['massnahme']['name']} {data['fiscal_year']['jahr']}"
    _style_title(c)
    headers = ["Nr", "Datum", "Auftraggeber", "Verwendungszweck", "Betrag (€)", "Kostenart", "Förderbetrag (€)", "Belege"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=2, column=i + 1, value=h)
        _style_colheader(cell)

    txs = data["transaktionen"]
    for idx, tx in enumerate(txs):
        r = idx + 3
        is_ausgabe = tx["betrag"] < 0
        values = [
            idx + 1, tx["datum"], tx["auftraggeber"] or "", tx["verwendungszweck"] or "",
            tx["betrag"], tx["kostenart"] or "", tx["betrag_foerderung"], tx["belege_count"],
        ]
        for ci, val in enumerate(values):
            cell = ws.cell(row=r, column=ci + 1, value=val)
            _style_data(cell, ci in (4, 6))
            if ci in (4, 6):
                _eur(cell)
                if ci == 4 and is_ausgabe:
                    cell.font = Font(size=10, name="Calibri", color="FFCC0000")

    if not txs:
        ws.merge_cells("A3:H3")
        c = ws["A3"]
        c.value = "Keine Transaktionen mit bestätigten Zuordnungen vorhanden."
        c.font = Font(italic=True, color="FF808080", size=10)
        c.alignment = Alignment(horizontal="center")


def _build_soll_ist(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Soll-Ist-Vergleich")
    for col, w in zip("ABCDE", [30, 16, 16, 16, 14]):
        ws.column_dimensions[col].width = w
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Soll-Ist-Vergleich — {data['massnahme']['name']} {data['fiscal_year']['jahr']}"
    _style_title(c)
    headers = ["Kostenart", "Bewilligt (€)", "Verausgabt (€)", "Differenz (€)", "Ausschöpfung %"]
    for i, h in enumerate(headers):
        _style_colheader(ws.cell(row=2, column=i + 1, value=h))

    positions = data["budget_positionen"]
    for idx, bp in enumerate(positions):
        r = idx + 3
        ausschoepfung = (bp["betrag_ist"] / bp["betrag_bewilligt"] * 100) if bp["betrag_bewilligt"] > 0 else 0
        is_ueber = bp["differenz"] < 0
        ampel = "FFFFC7CE" if ausschoepfung > 100 else ("FFFFEB9C" if ausschoepfung >= 90 else "FFE2EFDA")
        values = [bp["kostenart"], bp["betrag_bewilligt"], bp["betrag_ist"], bp["differenz"], ausschoepfung / 100]
        for ci, val in enumerate(values):
            cell = ws.cell(row=r, column=ci + 1, value=val)
            _style_data(cell, ci > 0)
            if 1 <= ci <= 3:
                _eur(cell)
                if ci == 3 and is_ueber:
                    cell.font = Font(size=10, name="Calibri", color="FFCC0000", bold=True)
            if ci == 4:
                cell.number_format = "0.0%"
                cell.fill = _fill(ampel)

    totals_row = len(positions) + 3
    tb = sum(bp["betrag_bewilligt"] for bp in positions)
    ti = sum(bp["betrag_ist"] for bp in positions)
    ta = (ti / tb) if tb > 0 else 0
    totals = ["GESAMT", tb, ti, tb - ti, ta]
    for ci, val in enumerate(totals):
        cell = ws.cell(row=totals_row, column=ci + 1, value=val)
        _style_sum(cell, ci > 0)
        if 1 <= ci <= 3:
            _eur(cell)
        if ci == 4:
            cell.number_format = "0.0%"


def generate_excel(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    wb.properties.creator = "FoerderFlow"
    _build_einnahmen_ausgaben(wb, data)
    _build_belegliste(wb, data)
    if data["budget_positionen"]:
        _build_soll_ist(wb, data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
