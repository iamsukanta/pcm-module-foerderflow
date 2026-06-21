"""Stundennachweis Excel — openpyxl port of lib/nachweis/stundennachweis-generator.ts.

One sheet per employee (sorted by name), monthly rows with VZÄ + AG-Brutto Projekt,
a Gesamt row, and confirmation/signature lines.
"""

from __future__ import annotations

import io
import re
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models.funding import FundingMeasure, FundingMeasureCostCenter
from app.models.payroll import Employee, MonthlyPayroll, PayrollAllocation
from app.services.vzae import berechne_vzae, berechne_vzae_anteil

MONTH_ABBR = ["Jan", "Feb", "Mär", "Apr", "Mai", "Jun", "Jul", "Aug", "Sep", "Okt", "Nov", "Dez"]
EUR_FMT = '#,##0.00 "€"'


def _fmt_monat(d: date) -> str:
    return f"{MONTH_ABBR[d.month - 1]} {d.year}"


def _fmt_date_de(d: date) -> str:
    return f"{d.day:02d}.{d.month:02d}.{d.year}"


def _sanitize_sheet(name: str) -> str:
    return re.sub(r"[:\\/?*\[\]]", "", name)[:31]


def _fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


def _style_title(cell) -> None:
    cell.font = Font(bold=True, size=13, color="FFFFFFFF", name="Calibri")
    cell.fill = _fill("FF1F3864")
    cell.alignment = Alignment(vertical="center", horizontal="left")


def _style_info(cell) -> None:
    cell.font = Font(size=10, color="FF404040", name="Calibri")
    cell.alignment = Alignment(vertical="center", horizontal="left")


def _style_colheader(cell) -> None:
    cell.font = Font(bold=True, size=10, color="FF1F3864", name="Calibri")
    cell.fill = _fill("FFDAE3F3")
    cell.alignment = Alignment(vertical="center", horizontal="center")
    cell.border = Border(bottom=Side(style="thin", color="FF2E75B6"))


def _style_data(cell, align: str = "left") -> None:
    cell.font = Font(size=10, name="Calibri")
    cell.alignment = Alignment(vertical="center", horizontal=align)


def _style_gesamt(cell, align: str = "left") -> None:
    cell.font = Font(bold=True, size=10, name="Calibri")
    cell.fill = _fill("FFFFE699")
    cell.alignment = Alignment(vertical="center", horizontal=align)
    cell.border = Border(
        top=Side(style="thin", color="FFBFBFBF"),
        bottom=Side(style="double", color="FF404040"),
    )


def _build_employee_sheet(wb: Workbook, emp: dict[str, Any], measure_name: str, von: date, bis: date) -> None:
    ws = wb.create_sheet(_sanitize_sheet(f"{emp['nachname']}, {emp['vorname']}"))
    for col, w in zip("ABCDE", [16, 22, 22, 10, 20]):
        ws.column_dimensions[col].width = w
    row = 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = f"Stundennachweis — {emp['vorname']} {emp['nachname']}"
    _style_title(c)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = f"Förderprojekt: {measure_name}"
    _style_info(c)
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = f"Zeitraum: {_fmt_date_de(von)} – {_fmt_date_de(bis)}"
    _style_info(c)
    row += 2

    headers = ["Monat", "Vertragliche Std/Woche", "Projektstunden/Woche", "VZÄ", "AG-Brutto Projekt"]
    for i, h in enumerate(headers):
        _style_colheader(ws.cell(row=row, column=i + 1, value=h))
    row += 1

    total_betrag = 0.0
    sum_vzae = 0.0
    for entry in emp["months"]:
        vzae_gesamt = berechne_vzae(entry["assigned_hours"], entry["standard_hours"])
        vzae_projekt = berechne_vzae_anteil(vzae_gesamt, entry["prozent_kst"])
        stunden_projekt = entry["assigned_hours"] * entry["prozent_kst"] / 100
        total_betrag += entry["betrag_anteil"]
        sum_vzae += vzae_projekt
        row_data = [
            _fmt_monat(entry["monat"]),
            round(entry["assigned_hours"], 2),
            round(stunden_projekt, 2),
            round(vzae_projekt, 2),
            entry["betrag_anteil"],
        ]
        for i, val in enumerate(row_data):
            cell = ws.cell(row=row, column=i + 1, value=val)
            align = "left" if i == 0 else ("center" if i == 3 else "right")
            _style_data(cell, align)
            if i == 4:
                cell.number_format = EUR_FMT
        row += 1

    avg_vzae = sum_vzae / len(emp["months"]) if emp["months"] else 0
    gesamt = ["Gesamt", "—", "—", round(avg_vzae, 2), total_betrag]
    for i, val in enumerate(gesamt):
        cell = ws.cell(row=row, column=i + 1, value=val)
        align = "left" if i == 0 else ("center" if i == 3 else "right")
        _style_gesamt(cell, align)
        if i == 4:
            cell.number_format = EUR_FMT
    row += 2

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = "Ich bestätige, dass die oben genannten Angaben korrekt sind."
    c.font = Font(size=10, italic=True, name="Calibri", color="FF404040")
    c.alignment = Alignment(vertical="center", horizontal="left")
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    c = ws[f"A{row}"]
    c.value = "Datum: _____________  Unterschrift: _____________"
    c.font = Font(size=10, name="Calibri", color="FF404040")
    c.alignment = Alignment(vertical="center", horizontal="left")


def generate_stundennachweis(
    db: Session, *, funding_measure_id: str, fiscal_year_id: str, org_id: str
) -> bytes:
    measure = db.execute(
        select(FundingMeasure).where(
            FundingMeasure.id == funding_measure_id, FundingMeasure.org_id == org_id
        )
    ).scalar_one_or_none()
    if measure is None:
        raise ValueError("Fördermassnahme nicht gefunden.")

    kst_ids = [
        r[0]
        for r in db.execute(
            select(FundingMeasureCostCenter.cost_center_id).where(
                FundingMeasureCostCenter.funding_measure_id == funding_measure_id,
                FundingMeasureCostCenter.org_id == org_id,
            )
        ).all()
    ]

    allocations = (
        db.execute(
            select(PayrollAllocation)
            .join(MonthlyPayroll, MonthlyPayroll.id == PayrollAllocation.payroll_id)
            .where(
                PayrollAllocation.org_id == org_id,
                PayrollAllocation.cost_center_id.in_(kst_ids) if kst_ids else False,
                MonthlyPayroll.fiscal_year_id == fiscal_year_id,
            )
            .order_by(MonthlyPayroll.monat.asc())
            .options(
                selectinload(PayrollAllocation.payroll).selectinload(MonthlyPayroll.employee)
            )
        )
        .scalars()
        .all()
    )

    by_employee: dict[str, dict[str, Any]] = {}
    for alloc in allocations:
        payroll = alloc.payroll
        emp: Employee = payroll.employee
        accum = by_employee.setdefault(
            emp.id,
            {
                "vorname": emp.vorname,
                "nachname": emp.nachname,
                "employee_code": emp.employee_code,
                "by_monat": {},
            },
        )
        key = payroll.monat.isoformat()
        existing = accum["by_monat"].get(key)
        if existing:
            existing["prozent_kst"] += float(alloc.prozent)
            existing["betrag_anteil"] += float(alloc.betrag_anteil)
        else:
            accum["by_monat"][key] = {
                "monat": payroll.monat,
                "assigned_hours": float(payroll.assigned_hours),
                "standard_hours": float(payroll.standard_hours),
                "prozent_kst": float(alloc.prozent),
                "betrag_anteil": float(alloc.betrag_anteil),
            }

    wb = Workbook()
    wb.remove(wb.active)
    wb.properties.creator = "FoerderFlow"

    sorted_emps = sorted(
        by_employee.values(), key=lambda e: f"{e['nachname']} {e['vorname']}".lower()
    )
    for emp in sorted_emps:
        months = sorted(emp["by_monat"].values(), key=lambda m: m["monat"])
        _build_employee_sheet(
            wb,
            {"vorname": emp["vorname"], "nachname": emp["nachname"], "employee_code": emp["employee_code"], "months": months},
            measure.name,
            measure.laufzeit_von,
            measure.laufzeit_bis,
        )

    if not sorted_emps:
        ws = wb.create_sheet("Keine Daten")
        ws["A1"].value = "Keine Abrechnungsdaten für diesen Zeitraum gefunden."

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
